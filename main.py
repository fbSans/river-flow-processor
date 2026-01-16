#/usr/bin/env python


import openpyxl
import xlrd
import os

XLS_DIR = './sheets'
OUT_DIR = './output'

XLS_SHEETS = [
    'OUTUBRO',
    'NOVEMBRO',
    'DEZEMBRO',
    'JANEIRO',
    'FEVEREIRO',
    'MARÇO',
    'ABRIL',
    'MAIO',
    'JUNHO',
    'JULHO',
    'AGOSTO',
    'SETEMBRO',
]

XLS_SHEET_DESC = {
        'OUTUBRO'  : { 'cdays' : 31, 'ldays': 31 },
        'NOVEMBRO' : { 'cdays' : 30, 'ldays': 30 },
        'DEZEMBRO' : { 'cdays' : 31, 'ldays': 31 },
        'JANEIRO'  : { 'cdays' : 31, 'ldays': 31 },
        'FEVEREIRO': { 'cdays' : 28, 'ldays': 29 },
        'MARÇO'    : { 'cdays' : 31, 'ldays': 31 },
        'ABRIL'    : { 'cdays' : 30, 'ldays': 30 },
        'MAIO'     : { 'cdays' : 31, 'ldays': 31 },
        'JUNHO'    : { 'cdays' : 30, 'ldays': 30 },
        'JULHO'    : { 'cdays' : 31, 'ldays': 31 },
        'AGOSTO'   : { 'cdays' : 31, 'ldays': 31 },
        'SETEMBRO' : { 'cdays' : 30, 'ldays': 30 },
}

#lines and cols are 1-based
def sheet_get_col(sheet, col, min_line, lines_count):
    res = []
    col -= 1
    for lin in range(min_line - 1, min_line + lines_count - 1):
        try:
            if sheet.cell(lin, col).ctype != xlrd.XL_CELL_NUMBER:
                res.append('f')
            else: 
                res.append(sheet.cell(lin, col).value)
        except IndexError as e :
            print(e, lin, col, sheet)

    assert len(res) >= 28, "Unexpected lenght, expected at least 28 entries"
    return res


def year_days_selector(year):
    if (year % 4 == 0 and year % 100 != 0 or year % 400 == 0): return 'ldays'
    return 'cdays'


def get_year_by_name(filename):
   return int(filename.split('.')[1].split('-')[0]) 


def get_sheet_avgs(sheet, day_count):
    START_LINE = 20
    AVGS_COL = 11
    return sheet_get_col(sheet, AVGS_COL, START_LINE, day_count)

def save_col_as_xlsx(name, data, col_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = col_name

    for i, item in enumerate(data, start=2):
        ws[f"A{i}"] = item
    wb.save(name)



if __name__ == '__main__':
    files = os.listdir(XLS_DIR)
    files.sort()
    data = []
    prev_year = None 
    seg_begin = None
    cached_file_name = None
    total_days = 0

    if not os.path.exists(OUT_DIR):
        os.makedirs(OUT_DIR)

    for filename in files:
        _, ext = os.path.splitext(filename)
        if ext != '.xls': 
            continue
        
        filepath = os.path.join(XLS_DIR, filename)
        year = get_year_by_name(filename)

        if seg_begin is None:
            seg_begin = year

        if prev_year is not None and prev_year + 1 != year:
            print("[INFO]\033[33m SEGMENT BREAK DETECTED.\033[0m")
            colname = filename.split('.')[0]
            savename = os.path.join(OUT_DIR, f"{filename.split('.')[0]}_{seg_begin}_processed.xlsx")
            print(f'[INFO] Saving data in {savename}...')
            save_col_as_xlsx(savename, data, colname)
            print(f"Saved {len(data)} items.")
            total_days += len(data)
            data = []
            seg_begin = year
        prev_year = year

        print(f"[INFO] Processing {filename}...")

        #Work on current
        wb = xlrd.open_workbook(filepath)
        days_selector = year_days_selector(year)


        sheets = [(wb.sheet_by_name(name), XLS_SHEET_DESC[name][days_selector]) for name in XLS_SHEETS]

        avgs = [
            get_sheet_avgs(sheet, day_count)
            for sheet, day_count in sheets
        ]

        for item in avgs:
            data += item

        print(f"[INFO] Done processing {filename}.")
        if not cached_file_name: cached_file_name = filename.split('.')[0]

    #handling remaing data
    colname = cached_file_name
    savename = os.path.join(OUT_DIR, f"{cached_file_name}_{seg_begin}_processed.xlsx")
    print(f'[INFO] Saving data in {savename}...')
    save_col_as_xlsx(savename, data, colname)
    print(f"Saved {len(data)} items.")
    total_days += len(data)
    print(f"Saved a total of {total_days} cells.")  

